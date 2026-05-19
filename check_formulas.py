import openpyxl
import re

file_path = r"C:\Work\1_Projects\202605_项目追踪表线上化\S520_金山中心_项目执行跟踪详细数据2026年05月.xlsx"
wb = openpyxl.load_workbook(file_path)
ws = wb.active

headers = {}
for cell in ws[4]: 
    if cell.value:
        col_idx = cell.column
        letter = openpyxl.utils.get_column_letter(col_idx)
        val = str(cell.value).replace('\n', '|')
        # Take the first part before '|' if exists, else whole
        headers[letter] = val.split('|')[0]

# Check row 5 (First data row) and row 6
# Previous output suggested row 5 might be data, let's check row 5 and 6 to be safe
for r in [5, 6]:
    if ws.cell(row=r, column=1).value is not None or ws.cell(row=r, column=4).value is not None: # Check for non-empty
        print(f"Checking Row {r} for formulas...")
        for cell in ws[r]:
            col_letter = openpyxl.utils.get_column_letter(cell.column)
            val = cell.value
            if isinstance(val, str) and val.startswith('='):
                # Found a formula
                header = headers.get(col_letter, 'Unknown')
                print(f"  {col_letter} ({header}): {val}")
